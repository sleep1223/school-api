import time
from school_api import SchoolClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from ruamel import yaml
from pathlib import Path

url = 'http://221.233.125.126:8096'
account = '你的学号'
password = '你的密码'

GdstApi = SchoolClient(url)  # 注册一个学校A
client_a = GdstApi.user_login(account, password, timeout=2)

info_data = client_a.get_info(timeout=5)
print('用户个人信息', info_data)

schedule_data = client_a.get_plan_schedule(xm=info_data['real_name'], timeout=5)
print('教学计划查询页面默认展示的课程', schedule_data)


def save():
    """保存各专业数据"""
    njs = client_a.plan_schedule.getGrades()
    xys = client_a.plan_schedule.getFacultys()
    # zys = client_a.plan_schedule.getSpecialtys()
    # xqs = client_a.plan_schedule.getTerms()
    # print(njs, xys, zys, xqs, sep='\n')

    data = {
        'name': '湖北师范大学文理学院',
        'url': url,
        'grades': njs,
        'facultys': xys,
        'specialtys': [],
    }

    for xy in xys:
        columns = client_a.get_plan_schedule(xm=info_data['real_name'], schedule_grade=info_data['grade'], schedule_faculty=xy['value'], schedule_specialty='',
                                             schedule_term='0', schedule_nature=0, timeout=5)
        zys = client_a.plan_schedule.getSpecialtys()

        if len(columns) > 0:
            for zy in zys:
                data['specialtys'].append({
                    'major': zy,
                    'faculty': xy,
                })
                print(xy['name'], xy['value'], zy['name'], zy['value'])
        else:
            # print('空专业', info_data['grade'],xy['value'])
            time.sleep(10)
        time.sleep(1)

    with open(Path(__file__).parent / 'data.yml', 'w', encoding='utf-8') as f:
        yaml.dump(data, f, Dumper=yaml.RoundTripDumper, encoding='utf-8', allow_unicode=True)


def load(major_year='2019-2022'):
    """爬取各专业所有学期课程表"""

    with open(Path(__file__).parent / 'data.yml', 'r', encoding='utf-8') as f:
        data = yaml.safe_load(f)

    for specialty in data['specialtys']:
        # print(specialty['faculty']['name'], specialty['faculty']['value'], specialty['major']['name'], specialty['major']['value'])
        now = time.strftime("%Y-%m-%d", time.localtime())
        pathfile = Path(__file__).parent / f'湖北师范大学文理学院 {now}.xlsx'
        if pathfile.is_file():
            wb = load_workbook(pathfile)
        else:
            wb = Workbook()
            wb.remove(wb['Sheet'])
        fontObj = Font(name=u'微软雅黑', bold=True, italic=False, size=12)
        align = Alignment(horizontal='center', vertical='center', wrap_text=False)

        if major_year:
            major_years = major_year.split('-')
            njs = reversed([str(x) for x in range(int(major_years[0]), int(major_years[1]) + 1)])
        else:
            njs = [x['value'] for x in client_a.plan_schedule.getGrades()]
            print('遍历所有年级课表', njs)

        for nj in njs:
            xy = specialty['faculty']['value']
            zy = specialty['major']['value']
            xym = specialty['faculty']['name']
            zym = specialty['major']['name']

            # 必须先按参数顺序获取一次，否则会出现获取不到数据的情况
            client_a.get_plan_schedule(xm=info_data['real_name'], schedule_grade=nj, schedule_faculty='', schedule_specialty='',
                                       schedule_term='0', schedule_nature=0, timeout=5)  # 全部学期 全部课程
            client_a.get_plan_schedule(xm=info_data['real_name'], schedule_grade=nj, schedule_faculty=xy, schedule_specialty='',
                                       schedule_term='0', schedule_nature=0, timeout=5)
            columns = client_a.get_plan_schedule(xm=info_data['real_name'], schedule_grade=nj, schedule_faculty=xy, schedule_specialty=zy,
                                                 schedule_term='0', schedule_nature=0, timeout=5)

            if len(columns) <= 1:  # 忽略空课表和非本专业课表
                # print('课表少于1行', nj, xym, zym)
                continue
            if zym not in wb.sheetnames:
                ws = wb.create_sheet(zym, 0)
            else:
                ws = wb[zym]
            ws.append(['', '', ''])
            ws.append(['', '', ''])
            ws.append([f'{nj}级 {xym} {zym}专业'])
            row_index = ws._current_row
            ws.merge_cells(f'A{row_index}:J{row_index}')
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['G'].width = 18
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 15

            # 课程代码	课程名称	学分	周学时	考核方式	课程性质	课程类别	建议修读学期	课程信息	辅修标识	专业方向	组代码	模块代码	通过情况	起始结束周	是否短学期

            for column in columns:
                ws.append(column)
            print('遍历完成', nj, xym, zym, len(columns) - 1)
            time.sleep(1)

        for worksheet in wb.worksheets:
            ws = worksheet
            ws.delete_cols(9, 6)

            for row in ws.rows:
                for sheet in row:
                    sheet.font = fontObj
                    sheet.alignment = align

        wb.save(pathfile)
        wb.close()
    print('课表遍历完成')


if __name__ == '__main__':
    # 判断各专业数据文件data.yml 是否存在
    path = Path(__file__).parent / 'data.yml'
    if not path.is_file():
        save()  # 保存各专业数据
    load()  # 爬取各专业所有学期课程表
